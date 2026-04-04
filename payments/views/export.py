from io import BytesIO

from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpRequest, HttpResponse
from openpyxl import Workbook

from ..models import (
    CheckoutSession,
    Company,
    CompanyUsageHistory,
    CreditHistory,
    CreditPlan,
    StripeCustomer,
    SubscriptionHistory,
    SubscriptionPlan,
    User,
)

MODELS = [
    ("Company", Company, ["id", "name", "created_at", "updated_at"]),
    ("User", User, ["id", "username", "company_id", "is_staff", "is_active", "date_joined"]),
    ("StripeCustomer", StripeCustomer, ["id", "company_id", "stripe_customer_id", "created_at", "updated_at"]),
    (
        "SubscriptionPlan",
        SubscriptionPlan,
        ["id", "name", "stripe_price_id", "monthly_document_limit", "monthly_ai_chat_limit", "created_at"],
    ),
    (
        "SubscriptionHistory",
        SubscriptionHistory,
        [
            "id", "stripe_customer_id", "subscription_plan_id",
            "stripe_subscription_id", "status", "current_period_start", "current_period_end", "created_at",
        ],
    ),
    ("CreditPlan", CreditPlan, ["id", "name", "stripe_price_id", "document_credits", "ai_chat_credits", "created_at"]),
    (
        "CreditHistory",
        CreditHistory,
        ["id", "stripe_customer_id", "credit_plan_id", "stripe_payment_id", "is_active", "created_at"],
    ),
    (
        "CompanyUsageHistory",
        CompanyUsageHistory,
        ["id", "company_id", "user_id", "type", "source", "created_at"],
    ),
    (
        "CheckoutSession",
        CheckoutSession,
        ["id", "stripe_customer_id", "stripe_session_id", "type", "status", "created_at"],
    ),
]


@staff_member_required
def export_all(request: HttpRequest) -> HttpResponse:
    """全テーブルをシートごとにまとめた Excel を返す."""
    wb = Workbook()
    wb.remove(wb.active)

    # Index シート
    index_ws = wb.create_sheet(title="Index")
    index_ws.append(["シート名", "テーブル名", "レコード数"])
    for sheet_name, model, _fields in MODELS:
        count = model.objects.count()
        index_ws.append([sheet_name, model._meta.db_table, count])  # noqa: SLF001
    for row_idx, (sheet_name, _model, _fields) in enumerate(MODELS, start=2):
        cell = index_ws.cell(row=row_idx, column=1)
        cell.hyperlink = f"#{sheet_name}!A1"
        cell.style = "Hyperlink"

    for sheet_name, model, fields in MODELS:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(fields)
        for obj in model.objects.all():
            ws.append([str(getattr(obj, f, "")) for f in fields])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="payment_core_export.xlsx"'
    return response
